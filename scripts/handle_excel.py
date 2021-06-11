from openpyxl import load_workbook
from scripts.handle_config import do_config
from scripts.handle_os import cases_path


class HandleExcel:

    def __init__(self, filename, sheetname=None):
        """
        excel的获取用例、写入封装
        :param filename: 文件名
        :param sheetname: 表单名
        """

        self.filename = filename
        self.sheetname = sheetname

    def get_cases(self):
        """
        获取所有用例方法
        :return: 嵌套字典的列表
        """
        wb = load_workbook(self.filename)
        if self.sheetname is None:
            ws = wb.active
        else:
            ws = wb[self.sheetname]
        head_data = tuple(ws.iter_rows(max_row=1, values_only=True))[0]
        data_lsit = []
        datas = tuple(ws.iter_rows(min_row=2,values_only=True))
        for data in datas:
            data_lsit.append(dict(zip(head_data, data)))
        return data_lsit

    def write_excel(self, row, actual, result):
        """
        写入excel操作
        :param row: 行数
        :param actual:实际结果
        :param result:用例执行结果
        :return:
        """
        other_wb = load_workbook(self.filename)
        if self.sheetname is None:
            other_ws = other_wb.active
        else:
            other_ws = other_wb[self.sheetname]
        if isinstance(row, int) and (2 <= row <= other_ws.max_row):
            other_ws.cell(row=row, column=do_config.get_int("file_path", "actual_column"), value=actual)
            other_ws.cell(row=row, column=do_config.get_int("file_path", "result_column"), value=result)
            other_wb.save(self.filename)
            other_wb.close()
        else:
            return "行号错误"


do_excel = HandleExcel(cases_path)
if __name__ == '__main__':
    excel=HandleExcel(cases_path)
    cases=excel.get_cases()
    print(cases)
    # a=excel.write_excel(2,"去","pass")


